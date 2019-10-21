import bs4
import openpyxl
import os
import re
import requests

from config import ERROR_TG_TOKEN, ADMIN_CHAT_ID, CURRENT_SHEET






def clean_text(text):
    text = text.strip()
    if '\xa0' in text:
        text = text.replace('\xa0', ' ')
    return text


def get_block_data(block):    
    fold_list = [
        'Местоположение', 'Адрес', 'Телефон', 'Эл. почта', 'Сайт', 'ИНН', 'Описание']
    items_list = block.find_all('li')
    block_data = []
    for fold in fold_list:        
        for item in items_list:
            try:
                fold_name = clean_text(
                    item.find('span', class_='dt').get_text()
                    )
                fold_value = clean_text(
                    item.find('span', class_='dd').get_text()
                    )
            except AttributeError:
                continue
            
            if fold == fold_name:
                block_data.append(fold_value)
                break
            # Если прошлись по всем строкам и не нашли, присваиваем значение по умолчанию
            if items_list.index(item) == (len(items_list) - 1):
                block_data.append('Не найдено поле ' + fold)

    return block_data


def write_html_file(html):
    with open('page.html', 'w') as f:
        f.write(html)

            
def write_data_to_excel(data_list):  
    pathh = os.path.join(os.getcwd(), 'objects.xlsx')
    if os.path.exists(pathh):
        wb = openpyxl.load_workbook(os.path.join(os.getcwd(), 'objects.xlsx'))
    else:
        wb = openpyxl.Workbook()
        wb.create_sheet(title = CURRENT_SHEET, index = 0)    
    
    sheet = wb[CURRENT_SHEET]
    start_row = sheet.max_row
    for i in range(start_row, start_row + len(data_list)):        
        sheet.append(data_list[i - start_row])
    
    wb.save('objects.xlsx')
    wb.close()
         

def get_cat_pages_count(html):
    soup = bs4.BeautifulSoup(html, features='lxml')
    pages_count = int(
        clean_text(soup.find('div', class_='paginator').find_all('a')[-2].get_text()))
    return pages_count


def get_cat_page_links(html):
    link_list = []
    soup = bs4.BeautifulSoup(html, features='lxml')
    raw_links_list = soup.find_all('td', class_='name')[1:]
    for link in raw_links_list:
        link = 'http://www.estateline.ru' + link.find('a').get('href')
        
        link_list.append(link)
    return link_list


def send_telegram_message(text):
    url = f'https://api.telegram.org/bot{ERROR_TG_TOKEN}/sendMessage'
    data = {'chat_id': ADMIN_CHAT_ID,
            'text': text
            }
    requests.get(url, params=data)



if __name__ == "__main__":
    pass

